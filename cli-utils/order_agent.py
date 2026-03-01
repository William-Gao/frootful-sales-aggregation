"""
Frootful Order Processing Agent

A Claude-powered agent that ingests raw order text, PDFs, images, or Excel files,
matches against pre-loaded customers/items, and creates proposals in staging.

Usage:
    python order_agent.py sample_orders/asta_tuesday.txt
    python order_agent.py order.pdf
    python order_agent.py order.xlsx
    python order_agent.py order.png
    python order_agent.py --text "Hey it's Cafe Sushi, 3 large basil and 2 small arugula for Friday"
    python order_agent.py --url https://example.com/order.jpg
    python order_agent.py --url https://example.com/order.pdf
"""

import json
import sys
import os
import base64
import mimetypes
from datetime import date

import anthropic
from dotenv import load_dotenv
from supabase import create_client

# Load env from agent/.env
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SECRET_KEY = os.environ["SUPABASE_SECRET_KEY"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
ORGANIZATION_ID = os.environ["ORGANIZATION_ID"]

supabase = create_client(SUPABASE_URL, SUPABASE_SECRET_KEY)
claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ─── File Processing ─────────────────────────────────────────────────────────

SUPPORTED_IMAGE_TYPES = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
SUPPORTED_DOC_TYPES = {".pdf"}
SUPPORTED_SPREADSHEET_TYPES = {".xlsx", ".xls", ".csv"}
SUPPORTED_TEXT_TYPES = {".txt", ".text", ".msg"}


def read_file(file_path: str) -> tuple[str, list[dict]]:
    """
    Read a file and return (text_content, content_blocks).

    - text_content: extracted text (for display)
    - content_blocks: list of Claude API content blocks to include in the message
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in SUPPORTED_TEXT_TYPES:
        return _read_text(file_path)
    elif ext in SUPPORTED_IMAGE_TYPES:
        return _read_image(file_path)
    elif ext in SUPPORTED_DOC_TYPES:
        return _read_pdf(file_path)
    elif ext in SUPPORTED_SPREADSHEET_TYPES:
        return _read_spreadsheet(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _read_text(file_path: str) -> tuple[str, list[dict]]:
    with open(file_path, "r") as f:
        text = f.read()
    return text, [{"type": "text", "text": text}]


def _read_image(file_path: str) -> tuple[str, list[dict]]:
    """Read image and return as base64 for Claude's vision."""
    mime_type = mimetypes.guess_type(file_path)[0] or "image/png"
    with open(file_path, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")
    return (
        f"[Image: {os.path.basename(file_path)}]",
        [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": mime_type,
                    "data": data,
                },
            },
            {
                "type": "text",
                "text": "This is an image of an order. Extract all order information from it.",
            },
        ],
    )


def _read_pdf(file_path: str) -> tuple[str, list[dict]]:
    """Read PDF and return as base64 for Claude's native PDF support."""
    with open(file_path, "rb") as f:
        data = base64.standard_b64encode(f.read()).decode("utf-8")
    return (
        f"[PDF: {os.path.basename(file_path)}]",
        [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": data,
                },
            },
            {
                "type": "text",
                "text": "This is a PDF order document. Extract all order information from it.",
            },
        ],
    )


def _read_spreadsheet(file_path: str) -> tuple[str, list[dict]]:
    """Read Excel/CSV and convert to text table for Claude."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        import csv
        with open(file_path, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
    else:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Install openpyxl to process Excel files: pip install openpyxl")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        rows = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows.append([f"--- Sheet: {sheet} ---"])
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        wb.close()

    # Format as text table
    lines = []
    for row in rows:
        lines.append(" | ".join(str(cell) for cell in row))
    text = "\n".join(lines)

    return (
        text,
        [
            {
                "type": "text",
                "text": f"This is spreadsheet data from {os.path.basename(file_path)}:\n\n{text}",
            }
        ],
    )


# ─── Load catalogs ───────────────────────────────────────────────────────────

# Lookup tables populated at startup — used by tool executors to enrich IDs with names
CUSTOMERS_BY_ID: dict[str, dict] = {}
ITEMS_BY_ID: dict[str, dict] = {}
VARIANTS_BY_ID: dict[str, dict] = {}


def load_customers() -> list[dict]:
    result = (
        supabase.table("customers")
        .select("id, name, email, phone, notes")
        .eq("active", True)
        .eq("organization_id", ORGANIZATION_ID)
        .order("name")
        .execute()
    )
    for c in result.data:
        CUSTOMERS_BY_ID[c["id"]] = c
    return result.data


def load_items() -> list[dict]:
    result = (
        supabase.table("items")
        .select("id, sku, name, description, item_variants(id, variant_code, variant_name, notes)")
        .eq("active", True)
        .eq("organization_id", ORGANIZATION_ID)
        .order("name")
        .execute()
    )
    for item in result.data:
        ITEMS_BY_ID[item["id"]] = item
        for v in item.get("item_variants", []):
            VARIANTS_BY_ID[v["id"]] = {**v, "item_id": item["id"], "item_name": item["name"]}
    return result.data


# ─── Tool Definitions ────────────────────────────────────────────────────────

NEW_ORDER_ITEM_SCHEMA = {
    "type": "object",
    "properties": {
        "item_id": {"type": "string"},
        "variant_id": {"type": "string"},
        "quantity": {"type": "number"},
    },
    "required": ["item_id", "variant_id", "quantity"],
}

MODIFY_ITEM_CHANGE_SCHEMA = {
    "type": "object",
    "description": (
        "A single item change. The 'type' field determines which other fields are expected:\n"
        "- add: requires item_id, variant_id, quantity\n"
        "- update: requires order_line_id, plus any fields being changed (variant_id, quantity)\n"
        "- remove: requires order_line_id only"
    ),
    "properties": {
        "type": {
            "type": "string",
            "enum": ["add", "update", "remove"],
        },
        "order_line_id": {
            "type": "string",
            "description": "Existing order_line ID from get_existing_orders. Required for update/remove.",
        },
        "item_id": {
            "type": "string",
            "description": "Item UUID. Required for add.",
        },
        "variant_id": {
            "type": "string",
            "description": "Variant UUID. Required for add. Optional for update (only if variant is changing).",
        },
        "quantity": {
            "type": "number",
            "description": "Required for add. Optional for update (only if quantity is changing).",
        },
    },
    "required": ["type"],
}

TOOLS = [
    {
        "name": "get_existing_orders",
        "description": (
            "Get a customer's upcoming orders (delivery_date >= today). Use this to "
            "determine if the incoming order is NEW or a CHANGE to an existing order. "
            "Returns order ID, delivery date, status, and current line items."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "customer_id": {"type": "string", "description": "The customer UUID"},
                "delivery_date": {
                    "type": "string",
                    "description": "Optional: filter to a specific date (YYYY-MM-DD)",
                },
            },
            "required": ["customer_id"],
        },
    },
    {
        "name": "create_new_order",
        "description": (
            "Create a proposal for a brand new order. Use this when NO existing order "
            "exists for this customer + delivery date."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "customer_id": {"type": "string"},
                "delivery_date": {"type": "string", "description": "YYYY-MM-DD"},
                "items": {"type": "array", "items": NEW_ORDER_ITEM_SCHEMA},
                "order_frequency": {
                    "type": "string",
                    "enum": ["one-time", "recurring"],
                    "description": "Whether this is a one-time order or a recurring/standing order",
                },
            },
            "required": ["customer_id", "delivery_date", "items", "order_frequency"],
        },
    },
    {
        "name": "modify_order",
        "description": (
            "Modify an existing order. Provide the order_id and a changes object "
            "describing what to change."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "order_id": {"type": "string", "description": "The existing order UUID"},
                "changes": {
                    "type": "object",
                    "description": "Fields to change. All are optional — only include what's changing.",
                    "properties": {
                        "customer_id": {
                            "type": "string",
                            "description": "New customer ID, if the order is being reassigned",
                        },
                        "delivery_date": {
                            "type": "string",
                            "description": "New delivery date (YYYY-MM-DD), if the date is changing",
                        },
                        "items": {
                            "type": "array",
                            "items": MODIFY_ITEM_CHANGE_SCHEMA,
                            "description": "Item-level changes (add/update/remove)",
                        },
                    },
                },
                "order_frequency": {
                    "type": "string",
                    "enum": ["one-time", "recurring"],
                    "description": "Whether this is a one-time change or a recurring/standing order change",
                },
            },
            "required": ["order_id", "changes", "order_frequency"],
        },
    },
    {
        "name": "cancel_order",
        "description": "Cancel an existing order entirely.",
        "input_schema": {
            "type": "object",
            "properties": {
                "order_id": {"type": "string", "description": "The existing order UUID"},
                "customer_id": {"type": "string"},
                "order_frequency": {
                    "type": "string",
                    "enum": ["one-time", "recurring"],
                    "description": "Whether this cancels a one-time order or a recurring/standing order",
                },
            },
            "required": ["order_id", "customer_id", "order_frequency"],
        },
    },
]


def build_system_prompt(customers: list[dict], items: list[dict]) -> str:
    # Format customers compactly
    customer_lines = []
    for c in customers:
        parts = [f"  {c['name']} (id: {c['id']})"]
        if c.get("email"):
            parts[0] += f" email: {c['email']}"
        if c.get("phone"):
            parts[0] += f" phone: {c['phone']}"
        if c.get("notes"):
            parts[0] += f" — {c['notes']}"
        customer_lines.append(parts[0])

    # Format items compactly
    item_lines = []
    for item in items:
        variants = item.get("item_variants", [])
        variant_str = ", ".join(
            f"{v['variant_code']}={v['variant_name']} (id:{v['id']})"
            for v in sorted(variants, key=lambda v: v["variant_code"])
        )
        line = f"  {item['name']} [SKU: {item['sku']}] (id: {item['id']}) → variants: {variant_str}"
        item_lines.append(line)

    return f"""You are Frootful's order processing agent for Boston Microgreens.
You receive orders from restaurant customers via text messages, emails, PDFs, images, or spreadsheets.

CUSTOMERS:
{chr(10).join(customer_lines)}

ITEMS & VARIANTS:
{chr(10).join(item_lines)}

YOUR WORKFLOW:
1. Read the order content (text, PDF, image, or spreadsheet)
2. Identify the customer (match against the customer list above)
3. Match each ordered item to the catalog above — use the exact item IDs and variant IDs
4. Check if an existing order already exists for the delivery date (use get_existing_orders)
5. Call the appropriate tool:
   - No existing order → create_new_order
   - Existing order + customer wants changes → modify_order
   - Existing order + customer wants to cancel → cancel_order

RULES:
- Order frequency: determine if the order is "recurring" or "one-time":
  - "weekly", "every week", "standing order", "recurring", "regular", "same as usual" → "recurring"
  - Otherwise → "one-time"
- Variants: S = Small, L = Large, T20 = Tray 20
  "small" → S, "large" → L, "tray" or "tray of" → T20
- If the customer doesn't specify a variant, default to S (Small)
- A single message may reference multiple delivery dates — call the tool separately for each
- For modify_order: pass order_id and a changes object. Only include what's changing:
  - changes.customer_id — if the order is being reassigned
  - changes.delivery_date — if the delivery date is changing
  - changes.items — array of item changes, each with a type:
    - type "add": new item → requires item_id, variant_id, quantity
    - type "update": changing an existing line → requires order_line_id, plus only the fields changing (variant_id, quantity)
    - type "remove": canceling a line → requires only order_line_id
- Today's date is {date.today().isoformat()}
- CRITICAL: All delivery dates MUST be in the future. When an order says "Tuesday" or "Friday", calculate the NEXT occurrence that is AFTER today. Example: if today is Saturday 2026-02-28 and the order says "Tuesday", that means Tuesday 2026-03-03 (NOT the past Tuesday 2026-02-24). Do NOT create orders for past dates. Do NOT comment on dates being in the past — just use the correct future date.

Be concise. Match, check existing orders, submit."""


# ─── Tool Execution ──────────────────────────────────────────────────────────


def execute_tool(name: str, tool_input: dict) -> dict:
    """Execute a tool call against Supabase and return the result."""

    if name == "get_existing_orders":
        return _exec_get_existing_orders(tool_input)
    elif name == "create_new_order":
        return _exec_create_new_order(tool_input)
    elif name == "modify_order":
        return _exec_modify_order(tool_input)
    elif name == "cancel_order":
        return _exec_cancel_order(tool_input)
    else:
        return {"error": f"Unknown tool: {name}"}


def _exec_get_existing_orders(params: dict) -> list:
    customer_id = params["customer_id"]
    today = date.today().isoformat()

    query = (
        supabase.table("orders")
        .select("id, delivery_date, status, order_lines(id, line_number, item_id, quantity, status, item_variant_id, items(name, sku))")
        .eq("customer_id", customer_id)
        .eq("organization_id", ORGANIZATION_ID)
        .neq("status", "cancelled")
        .gte("delivery_date", today)
        .order("delivery_date")
        .limit(5)
    )
    if params.get("delivery_date"):
        query = query.eq("delivery_date", params["delivery_date"])

    return query.execute().data


def _resolve_customer(customer_id: str) -> str:
    """Look up customer name from ID."""
    c = CUSTOMERS_BY_ID.get(customer_id)
    return c["name"] if c else "Unknown"


def _resolve_item(item_id: str, variant_id: str) -> tuple[str, str]:
    """Look up item_name and variant_code from IDs."""
    item = ITEMS_BY_ID.get(item_id)
    variant = VARIANTS_BY_ID.get(variant_id)
    item_name = item["name"] if item else "Unknown"
    variant_code = variant["variant_code"] if variant else "?"
    return item_name, variant_code


def _exec_create_new_order(params: dict) -> dict:
    """Create a new_order proposal. All items become 'add' lines."""
    customer_id = params["customer_id"]
    customer_name = _resolve_customer(customer_id)

    proposal = _insert_proposal(
        proposal_type="new_order",
        order_id=None,
        order_frequency=params.get("order_frequency", "one-time"),
    )
    proposal_id = proposal["id"]

    lines_created = []
    for i, item in enumerate(params.get("items", [])):
        item_name, variant_code = _resolve_item(item["item_id"], item["variant_id"])
        line = _insert_proposal_line(
            proposal_id=proposal_id,
            line_number=i + 1,
            item_id=item["item_id"],
            item_name=item_name,
            variant_id=item["variant_id"],
            variant_code=variant_code,
            quantity=item["quantity"],
            change_type="add",
            order_line_id=None,
            delivery_date=params["delivery_date"],
            customer_id=customer_id,
            customer_name=customer_name,
        )
        lines_created.append(line)

    return {
        "proposal_id": proposal_id,
        "type": "new_order",
        "lines_created": len(lines_created),
        "delivery_date": params["delivery_date"],
        "customer_name": customer_name,
    }


def _exec_modify_order(params: dict) -> dict:
    """
    Modify an existing order. The 'changes' object can contain:
    - customer_id: reassign order to a different customer
    - delivery_date: change the delivery date
    - items: array of item-level changes (add/update/remove)
    """
    order_id = params["order_id"]
    changes = params.get("changes", {})

    # Look up existing order to get current customer_id and delivery_date
    existing_order = (
        supabase.table("orders")
        .select("customer_id, delivery_date")
        .eq("id", order_id)
        .single()
        .execute()
    ).data

    # Use changed values or fall back to existing
    customer_id = changes.get("customer_id") or existing_order["customer_id"]
    delivery_date = changes.get("delivery_date") or existing_order["delivery_date"]
    customer_name = _resolve_customer(customer_id)

    proposal = _insert_proposal(
        proposal_type="change_order",
        order_id=order_id,
        order_frequency=params.get("order_frequency", "one-time"),
    )
    proposal_id = proposal["id"]

    CHANGE_MAP = {"add": "add", "update": "modify", "remove": "remove"}

    lines_created = []
    for i, item_change in enumerate(changes.get("items", [])):
        change_type_raw = item_change.get("type", "add")
        change_type = CHANGE_MAP.get(change_type_raw, "add")
        order_line_id = item_change.get("order_line_id") or None

        if change_type_raw == "add":
            # Add: requires item_id, variant_id, quantity
            item_id = item_change["item_id"]
            variant_id = item_change["variant_id"]
            quantity = item_change["quantity"]
            item_name, variant_code = _resolve_item(item_id, variant_id)

        elif change_type_raw == "update":
            # Update: requires order_line_id, optional variant_id + quantity
            # Look up current line values to fill in unchanged fields
            existing_line = (
                supabase.table("order_lines")
                .select("item_id, item_variant_id, quantity")
                .eq("id", order_line_id)
                .single()
                .execute()
            ).data

            item_id = item_change.get("item_id") or existing_line["item_id"]
            variant_id = item_change.get("variant_id") or existing_line["item_variant_id"]
            quantity = item_change.get("quantity", existing_line["quantity"])
            item_name, variant_code = _resolve_item(item_id, variant_id)

        elif change_type_raw == "remove":
            # Remove: just needs order_line_id — look up the rest for display
            existing_line = (
                supabase.table("order_lines")
                .select("item_id, item_variant_id, quantity")
                .eq("id", order_line_id)
                .single()
                .execute()
            ).data

            item_id = existing_line["item_id"]
            variant_id = existing_line["item_variant_id"]
            quantity = existing_line["quantity"]
            item_name, variant_code = _resolve_item(item_id, variant_id)

        else:
            continue

        line = _insert_proposal_line(
            proposal_id=proposal_id,
            line_number=i + 1,
            item_id=item_id,
            item_name=item_name,
            variant_id=variant_id,
            variant_code=variant_code,
            quantity=quantity,
            change_type=change_type,
            order_line_id=order_line_id,
            delivery_date=delivery_date,
            customer_id=customer_id,
            customer_name=customer_name,
        )
        lines_created.append(line)

    # Mark existing order as pending review
    supabase.table("orders").update({"status": "pending_review"}).eq(
        "id", order_id
    ).execute()

    supabase.table("order_events").insert({
        "order_id": order_id,
        "type": "change_proposed",
        "metadata": {"proposal_id": proposal_id, "source": "agent"},
    }).execute()

    return {
        "proposal_id": proposal_id,
        "type": "change_order",
        "lines_created": len(lines_created),
        "delivery_date": delivery_date,
        "customer_name": customer_name,
    }


def _exec_cancel_order(params: dict) -> dict:
    """Cancel an existing order."""
    order_id = params["order_id"]
    customer_id = params["customer_id"]
    customer_name = _resolve_customer(customer_id)

    proposal = _insert_proposal(
        proposal_type="cancel_order",
        order_id=order_id,
        order_frequency=params.get("order_frequency", "one-time"),
    )
    proposal_id = proposal["id"]

    supabase.table("orders").update({"status": "pending_review"}).eq(
        "id", order_id
    ).execute()

    supabase.table("order_events").insert({
        "order_id": order_id,
        "type": "cancel_proposed",
        "metadata": {"proposal_id": proposal_id, "source": "agent"},
    }).execute()

    return {
        "proposal_id": proposal_id,
        "type": "cancel_order",
        "customer_name": customer_name,
    }


# ─── DB Helpers ─────────────────────────────────────────────────────────────


def _insert_proposal(proposal_type: str, order_id: str | None, order_frequency: str = "one-time") -> dict:
    """Insert a proposal row and return it."""
    result = (
        supabase.table("order_change_proposals")
        .insert({
            "organization_id": ORGANIZATION_ID,
            "order_id": order_id,
            "status": "pending",
            "type": proposal_type,
            "tags": {"source": "agent", "agent_version": "0.3", "order_frequency": order_frequency},
        })
        .execute()
    )
    return result.data[0]


def _insert_proposal_line(
    proposal_id: str,
    line_number: int,
    item_id: str,
    item_name: str,
    variant_id: str,
    variant_code: str,
    quantity: float,
    change_type: str,
    order_line_id: str | None,
    delivery_date: str,
    customer_id: str,
    customer_name: str,
) -> dict:
    """Insert a proposal line row and return it."""
    result = (
        supabase.table("order_change_proposal_lines")
        .insert({
            "proposal_id": proposal_id,
            "line_number": line_number,
            "item_id": item_id,
            "item_name": item_name,
            "item_variant_id": variant_id,
            "change_type": change_type,
            "order_line_id": order_line_id,
            "proposed_values": {
                "quantity": quantity,
                "variant_code": variant_code,
                "delivery_date": delivery_date,
                "customer_id": customer_id,
                "customer_name": customer_name,
                "organization_id": ORGANIZATION_ID,
            },
        })
        .execute()
    )
    return result.data[0]


# ─── Agent Loop ──────────────────────────────────────────────────────────────


def process_order(message_text: str, content_blocks: list[dict] | None = None) -> dict:
    """Run the agent loop: send message to Claude, execute tools, repeat until done."""

    print(f"\n{'='*60}")
    print("FROOTFUL ORDER AGENT")
    print(f"{'='*60}")

    print("\nLoading catalogs from staging...")
    customers = load_customers()
    items = load_items()
    print(f"  Loaded {len(customers)} customers, {len(items)} items")

    system_prompt = build_system_prompt(customers, items)

    # Build the user message content
    if content_blocks:
        # File-based input: use the content blocks (image/pdf/text)
        user_content = content_blocks + [
            {
                "type": "text",
                "text": f"\nProcess this order. Today's date: {date.today().isoformat()}",
            }
        ]
    else:
        # Plain text input
        user_content = (
            f"Process this incoming order message:\n\n"
            f"{message_text}\n\n"
            f"Today's date: {date.today().isoformat()}"
        )

    messages = [{"role": "user", "content": user_content}]
    turn = 0
    max_turns = 100

    print(f"\nInput:\n{message_text.strip()}")
    print(f"\n{'─'*60}")

    while turn < max_turns:
        turn += 1
        print(f"\n▶ Turn {turn}")

        response = claude.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=4096,
            system=system_prompt,
            tools=TOOLS,
            messages=messages,
        )

        messages.append({"role": "assistant", "content": response.content})

        # Print Claude's reasoning
        for block in response.content:
            if block.type == "text" and block.text.strip():
                print(f"  💬 {block.text}")

        if response.stop_reason == "end_turn":
            print(f"\n{'='*60}")
            print("✅ Agent finished")
            print(f"   Turns: {turn}")
            print(f"   Input tokens: {response.usage.input_tokens}")
            print(f"   Output tokens: {response.usage.output_tokens}")
            print(f"{'='*60}")
            return {"success": True, "turns": turn}

        # Execute tool calls
        tool_results = []
        for block in response.content:
            if block.type == "tool_use":
                tool_name = block.name
                tool_input = block.input
                print(f"  🔧 {tool_name}({json.dumps(tool_input, indent=None)[:120]})")

                try:
                    result = execute_tool(tool_name, tool_input)
                    result_str = json.dumps(result, default=str)
                    display = result_str[:200] + "..." if len(result_str) > 200 else result_str
                    print(f"  📦 → {display}")
                except Exception as e:
                    result_str = json.dumps({"error": str(e)})
                    print(f"  ❌ Error: {e}")

                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result_str,
                })

        messages.append({"role": "user", "content": tool_results})

    print("\n⚠️  Max turns reached")
    return {"success": False, "error": "max_turns_reached", "turns": turn}


# ─── CLI ─────────────────────────────────────────────────────────────────────


def _url_content_blocks(url: str) -> tuple[str, list[dict]]:
    """Build Claude API content blocks from a URL (image or PDF)."""
    lower = url.lower().split("?")[0]  # strip query params for extension check

    if any(lower.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".gif", ".webp")):
        return (
            f"[Image URL: {url}]",
            [
                {
                    "type": "image",
                    "source": {"type": "url", "url": url},
                },
                {
                    "type": "text",
                    "text": "This is an image of an order. Extract all order information from it.",
                },
            ],
        )
    elif lower.endswith(".pdf"):
        return (
            f"[PDF URL: {url}]",
            [
                {
                    "type": "document",
                    "source": {"type": "url", "url": url},
                },
                {
                    "type": "text",
                    "text": "This is a PDF order document. Extract all order information from it.",
                },
            ],
        )
    else:
        # Default to image — Claude will error if it's not valid
        return (
            f"[URL: {url}]",
            [
                {
                    "type": "image",
                    "source": {"type": "url", "url": url},
                },
                {
                    "type": "text",
                    "text": "This is an order document. Extract all order information from it.",
                },
            ],
        )


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python order_agent.py <file>          # .txt, .pdf, .png, .jpg, .xlsx, .csv")
        print('  python order_agent.py --text "order message here"')
        print('  python order_agent.py --url <image-or-pdf-url>')
        sys.exit(1)

    if sys.argv[1] == "--text":
        message_text = " ".join(sys.argv[2:])
        result = process_order(message_text)
    elif sys.argv[1] == "--url":
        if len(sys.argv) < 3:
            print("Error: --url requires a URL argument")
            sys.exit(1)
        url = sys.argv[2]
        print(f"Using URL source: {url}")
        text, content_blocks = _url_content_blocks(url)
        result = process_order(text, content_blocks)
    else:
        file_path = sys.argv[1]
        if not os.path.isabs(file_path):
            file_path = os.path.join(os.path.dirname(__file__), file_path)

        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            sys.exit(1)

        print(f"Reading file: {file_path}")
        text, content_blocks = read_file(file_path)
        result = process_order(text, content_blocks)

    if result["success"]:
        print("\nDone! Check the staging dashboard to see the proposal.")
    else:
        print(f"\nFailed: {result.get('error')}")


if __name__ == "__main__":
    main()
